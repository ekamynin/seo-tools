import io
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from ahrefs_api import fetch_referring_domains
from cache import fetch_full_catalog

st.set_page_config(
    page_title="Backlink Gap",
    page_icon="🎯",
    layout="wide",
)

AHREFS_KEY = st.secrets.get("AHREFS_API_KEY", "")
COLLAB_KEY = st.secrets.get("COLLABORATOR_API_KEY", "")


def _norm(raw: str) -> str:
    """Strip protocol, www, paths — return bare domain."""
    d = raw.strip().lower()
    d = re.sub(r"^https?://", "", d)
    d = re.sub(r"^www\.", "", d)
    return d.split("/")[0].split("?")[0]


def _fetch_site_donors(api_key: str, domain: str, limit: int) -> dict[str, dict]:
    """Fetch referring domains. Returns {norm_domain: {dr, traffic, dofollow, raw_domain}}."""
    items = fetch_referring_domains(api_key, domain, limit=limit)
    result = {}
    for item in items:
        d = _norm(item.get("domain", ""))
        if d:
            result[d] = {
                "dr": item.get("domain_rating"),
                "traffic": item.get("traffic_domain"),
                "dofollow": (item.get("dofollow_links") or 0) > 0,
                "raw_domain": item.get("domain", d),
            }
    return result


# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ Налаштування")
    st.divider()
    donor_limit = st.select_slider(
        "Донорів на сайт",
        options=[500, 1000, 2000, 3000, 5000],
        value=1000,
        help="Топ N донорів за DR, що завантажуються для кожного сайту з Ahrefs.",
    )
    min_dr = st.number_input(
        "Мінімальний DR донора",
        min_value=0,
        max_value=100,
        value=0,
        help="Показувати тільки донорів з DR ≥ цього значення.",
    )
    min_traffic = st.number_input(
        "Мінімальний трафік донора",
        min_value=0,
        value=0,
        step=1000,
        help="Показувати тільки донорів з органічним трафіком ≥ цього значення.",
    )
    collab_only = st.toggle(
        "Тільки з Collaborator",
        value=False,
        help="Показати в Gap тільки донорів, що є в каталозі Collaborator.",
    )
    st.divider()
    st.caption("Backlink Gap v1.1")


# ── Main UI ───────────────────────────────────────────────────────────────────
st.title("🎯 Backlink Gap")
st.caption("Пошук донорів, які лінкують на конкурентів, але ще не на тебе.")

if not AHREFS_KEY:
    st.error("❌ Не знайдено AHREFS_API_KEY у секретах. Додай його в Settings → Secrets.")
    st.stop()

# ── Inputs ────────────────────────────────────────────────────────────────────
col_my, col_c1, col_c2, col_c3 = st.columns(4)
with col_my:
    my_site = st.text_input("🏠 Мій сайт", placeholder="mysite.com.ua")
with col_c1:
    comp1 = st.text_input("⚔️ Конкурент 1", placeholder="competitor1.ua")
with col_c2:
    comp2 = st.text_input("⚔️ Конкурент 2", placeholder="competitor2.ua")
with col_c3:
    comp3 = st.text_input("⚔️ Конкурент 3", placeholder="competitor3.ua")

competitors_raw = [c.strip() for c in [comp1, comp2, comp3] if c.strip()]
run_disabled = not my_site.strip() or not competitors_raw

if run_disabled and (my_site.strip() or any(c.strip() for c in [comp1, comp2, comp3])):
    st.caption("ℹ️ Вкажи мій сайт та хоча б одного конкурента.")

run = st.button(
    "🔍 Аналізувати",
    type="primary",
    disabled=run_disabled,
    use_container_width=True,
)

if run:
    my_domain = _norm(my_site.strip())
    comp_domains = [_norm(c) for c in competitors_raw]
    all_targets = [my_domain] + comp_domains

    # ── Fetch donors from Ahrefs ──────────────────────────────────────────
    donors_by_site: dict[str, dict] = {}
    progress = st.progress(0.0, text="Завантажуємо донорів з Ahrefs…")
    fetch_errors = []

    with ThreadPoolExecutor(max_workers=4) as executor:
        futures = {
            executor.submit(_fetch_site_donors, AHREFS_KEY, t, donor_limit): t
            for t in all_targets
        }
        done_count = 0
        for future in as_completed(futures):
            target = futures[future]
            try:
                donors_by_site[target] = future.result()
            except Exception as e:
                donors_by_site[target] = {}
                fetch_errors.append(f"{target}: {e}")
            done_count += 1
            progress.progress(
                done_count / len(all_targets),
                text=f"Завантажено {done_count}/{len(all_targets)} сайтів…",
            )

    progress.empty()
    for err in fetch_errors:
        st.warning(f"⚠️ Помилка при завантаженні {err}")

    my_donors = donors_by_site.get(my_domain, {})

    # ── Apply filters ─────────────────────────────────────────────────────
    def _passes_filters(v: dict) -> bool:
        if min_dr > 0 and (v.get("dr") or 0) < min_dr:
            return False
        if min_traffic > 0 and (v.get("traffic") or 0) < min_traffic:
            return False
        return True

    my_donors = {d: v for d, v in my_donors.items() if _passes_filters(v)}
    for cd in comp_domains:
        donors_by_site[cd] = {
            d: v for d, v in donors_by_site.get(cd, {}).items()
            if _passes_filters(v)
        }

    # Build combined competitor donor pool (highest DR wins for duplicates)
    all_comp_donors: dict[str, dict] = {}
    for cd in comp_domains:
        for d, v in donors_by_site.get(cd, {}).items():
            existing_dr = (all_comp_donors.get(d) or {}).get("dr") or 0
            if d not in all_comp_donors or (v.get("dr") or 0) > existing_dr:
                all_comp_donors[d] = v

    my_set = set(my_donors.keys())
    comp_set = set(all_comp_donors.keys())
    gap_set = comp_set - my_set
    shared_set = my_set & comp_set

    # ── Collaborator enrichment ───────────────────────────────────────────
    collab_lookup: dict = {}
    if COLLAB_KEY:
        with st.spinner("Звіряємо з каталогом Collaborator…"):
            try:
                df_catalog = fetch_full_catalog(COLLAB_KEY)
                collab_lookup = {
                    _norm(row["domain"]): row
                    for _, row in df_catalog.iterrows()
                    if row.get("domain")
                }
            except Exception as e:
                st.warning(f"⚠️ Помилка завантаження Collaborator: {e}")

    def _collab(domain: str) -> tuple:
        info = collab_lookup.get(domain)
        if info is None:
            return False, None, None, ""
        price = int(info["price"]) if info.get("price") else None
        price_w = int(info["price_writing"]) if info.get("price_writing") else None
        return True, price, price_w, info.get("collaborator_url", "")

    def _which_comps(domain: str) -> str:
        return ", ".join(cd for cd in comp_domains if domain in donors_by_site.get(cd, {}))

    def _comp_count(domain: str) -> int:
        return sum(1 for cd in comp_domains if domain in donors_by_site.get(cd, {}))

    # ── Gap DataFrame ─────────────────────────────────────────────────────
    gap_rows = []
    for d in gap_set:
        info = all_comp_donors.get(d, {})
        in_c, price, price_w, c_url = _collab(d)
        gap_rows.append({
            "Домен": info.get("raw_domain", d),
            "DR": info.get("dr"),
            "Трафік": info.get("traffic"),
            "Dofollow": "✅" if info.get("dofollow") else "—",
            "К-сть конкурентів": _comp_count(d),
            "Конкуренти": _which_comps(d),
            "В Collaborator": "Так" if in_c else "Ні",
            "Ціна публікації (грн)": price,
            "Ціна написання (грн)": price_w,
            "Collaborator": c_url,
        })
    df_gap = pd.DataFrame(gap_rows)
    if not df_gap.empty:
        df_gap = df_gap.sort_values(
            ["К-сть конкурентів", "DR"], ascending=[False, False]
        ).reset_index(drop=True)

    if collab_only and not df_gap.empty:
        df_gap = df_gap[df_gap["В Collaborator"] == "Так"].reset_index(drop=True)

    # ── Helper: повний список донорів для одного сайту (для Excel) ────────
    def _site_df(donors: dict) -> pd.DataFrame:
        rows = []
        for d, info in donors.items():
            in_c, price, price_w, c_url = _collab(d)
            rows.append({
                "Домен": info.get("raw_domain", d),
                "DR": info.get("dr"),
                "Трафік": info.get("traffic"),
                "Dofollow": "✅" if info.get("dofollow") else "—",
                "В Collaborator": "Так" if in_c else "Ні",
                "Ціна публікації (грн)": price,
                "Ціна написання (грн)": price_w,
                "Collaborator": c_url,
            })
        df = pd.DataFrame(rows)
        if not df.empty:
            df = df.sort_values("DR", ascending=False).reset_index(drop=True)
        return df

    # ── Summary metrics ───────────────────────────────────────────────────
    st.divider()
    in_collab_count = int(df_gap["В Collaborator"].eq("Так").sum()) if not df_gap.empty else 0
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("🎯 Gap донорів", len(gap_set))
    c2.metric("📋 З них в Collaborator", in_collab_count)
    c3.metric("🤝 Спільних донорів", len(shared_set))
    c4.metric("🔗 Моїх донорів", len(my_set))

    # ── Gap таблиця на екрані ─────────────────────────────────────────────
    COL_CFG = {
        "DR": st.column_config.NumberColumn(format="%d"),
        "Трафік": st.column_config.NumberColumn(format="%d"),
        "Ціна публікації (грн)": st.column_config.NumberColumn(format="%d"),
        "Ціна написання (грн)": st.column_config.NumberColumn(format="%d"),
        "Collaborator": st.column_config.LinkColumn(
            "Collaborator",
            display_text="Відкрити",
            help="Відкрити на Collaborator.pro",
        ),
    }

    st.caption("Донори конкурентів, яких у тебе ще немає. Відсортовано: к-сть конкурентів → DR.")
    if df_gap.empty:
        st.info("ℹ️ Gap-донорів не знайдено.")
    else:
        gap_cols = [
            "Домен", "DR", "Трафік", "Dofollow", "К-сть конкурентів", "Конкуренти",
            "В Collaborator", "Ціна публікації (грн)", "Ціна написання (грн)", "Collaborator",
        ]
        display_df = df_gap[[c for c in gap_cols if c in df_gap.columns]]

        def _style_collab(df):
            styles = pd.DataFrame("", index=df.index, columns=df.columns)
            styles.loc[df["В Collaborator"] == "Так"] = "background-color: #1a3a20"
            return styles

        st.dataframe(
            display_df.style.apply(_style_collab, axis=None),
            use_container_width=True,
            hide_index=True,
            column_config=COL_CFG,
        )

    # ── Excel — повна вигрузка по кожному сайту ───────────────────────────
    excel_sheets: dict[str, pd.DataFrame] = {}

    if not df_gap.empty:
        excel_sheets["Gap"] = df_gap

    df_my_xl = _site_df(my_donors)
    if not df_my_xl.empty:
        excel_sheets[f"Мій сайт ({my_domain})"[:31]] = df_my_xl

    for cd in comp_domains:
        df_comp = _site_df(donors_by_site.get(cd, {}))
        if not df_comp.empty:
            excel_sheets[f"Конкурент ({cd})"[:31]] = df_comp

    if excel_sheets:
        st.divider()
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            for sheet_name, df in excel_sheets.items():
                df.to_excel(writer, index=False, sheet_name=sheet_name)
            # Clickable hyperlinks in Collaborator column
            wb = writer.book
            for sheet_name, df in excel_sheets.items():
                if "Collaborator" not in df.columns:
                    continue
                ws = wb[sheet_name]
                col_letter = get_column_letter(list(df.columns).index("Collaborator") + 1)
                for row_idx, url in enumerate(df["Collaborator"], start=2):
                    if url:
                        cell = ws[f"{col_letter}{row_idx}"]
                        cell.value = "Відкрити"
                        cell.hyperlink = url
                        cell.font = Font(color="0563C1", underline="single")
        buf.seek(0)
        st.download_button(
            "📥 Завантажити Excel",
            data=buf,
            file_name=f"backlink_gap_{my_domain}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
